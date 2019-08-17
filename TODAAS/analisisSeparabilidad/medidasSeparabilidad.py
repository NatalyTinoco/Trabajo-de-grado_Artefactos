# -*- coding: utf-8 -*-
"""
Created on Wed Aug 14 22:19:30 2019

@author: Nataly
"""


## Import the packages
import numpy as np
from scipy import stats

def tstudent(a,b,N):
#    ## Define 2 random distributions
#    #Sample Size
#    N = 10
#    #Gaussian distributed data with mean = 2 and var = 1
#    a = np.random.randn(N) + 2
#    #Gaussian distributed data with with mean = 0 and var = 1
#    b = np.random.randn(N)
#    ## Calculate the Standard Deviation
    #Calculate the variance to get the standard deviation
    
    #For unbiased max likelihood estimate we have to divide the var by N-1, and therefore the parameter ddof = 1
    var_a = a.var(ddof=1)
    var_b = b.var(ddof=1)
    
    #std deviation
    s = np.sqrt((var_a + var_b)/2)   
    
    ## Calculate the t-statistics
    t = (a.mean() - b.mean())/(s*np.sqrt(2/N)) 
    ## Compare with the critical t-value
    #Degrees of freedom
    df = 2*N - 2
    
    #p-value after comparison with the t 
    p = 1 - stats.t.cdf(t,df=df)
    print("t = " + str(t))
    print("p = " + str(2*p))
    ### You can see that after comparing the t statistic with the critical t value (computed internally) we get a good p value of 0.0005 and thus we reject the null hypothesis and thus it proves that the mean of the two distributions are different and statistically significant.
    ## Cross Checking with the internal scipy function
    t2, p2 = stats.ttest_ind(a,b)
    print("t = " + str(t2))
    print("p = " + str(p2))
    return []
def tstudent2(a,b):   
    from scipy.stats import ttest_rel
    stat, p = ttest_rel(a, b)
    return stat,p


import openpyxl
doc = openpyxl.load_workbook('CaracteristicasValoresSingulares500x500.xlsx')
doc.get_sheet_names()
hoja = doc.get_sheet_by_name('Sheet1')
hoja.title
hoja2 = doc.get_sheet_by_name('Hoja1')
hoja2.title

#beta dm
hoja.cell(row=1,column=1).value
betaDM=[]
for i in range (3,354):
    betaDM.append(hoja.cell(row=i,column=1).value)

        
hoja.cell(row=1,column=9).value
betaRE=[]
for i in range (3,474):
    betaRE.append(hoja.cell(row=i,column=9).value)


hoja.cell(row=1,column=18).value
betaNO=[]
for i in range (3,474):
    betaNO.append(hoja.cell(row=i,column=18).value)


tstudent(np.asarray(betaDM),np.asarray(betaRE[0:len(betaDM)]),len(betaDM))
stat,p=tstudent2(np.asarray(betaDM),np.asarray(betaRE[0:len(betaDM)]))
print('stat',stat)
print('p',p)
#
#N=10
#a = np.random.randn(N) + 2
#b = np.random.randn(N)
#tstudent(np.asarray(a),np.asarray(b),N)
#stat,p=tstudent2(np.asarray(b),np.asarray(a))
#print('stat',stat)
#print('p',p)